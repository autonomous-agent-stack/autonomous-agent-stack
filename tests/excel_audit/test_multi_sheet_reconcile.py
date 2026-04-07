"""Golden test group 5: Multi-sheet reconciliation.

Tests cross-sheet joins, multiple workbooks, and report generation.
"""
from __future__ import annotations

from pathlib import Path

import pytest

from excel_audit.contracts import ExcelAuditRule, RuleDsl, SheetMapping
from excel_audit.parser import parse_workbook
from excel_audit.report import generate_report
from excel_audit.reconcile import reconcile
from excel_audit.workbook_runner import run_audit
from excel_audit.contracts import AuditReport, AuditResult, AuditFinding, AuditSeverity

pytest.importorskip("openpyxl")
from openpyxl import Workbook


@pytest.fixture()
def multi_dir(tmp_path: Path) -> Path:
    d = tmp_path / "multi"
    d.mkdir()

    # Single workbook with both source and target sheets
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "销售明细"
    ws1.append(["订单编号", "订单状态", "销售额", "提成比例"])
    ws1.append(["MS001", "已完成", 10000, 0.05])
    ws1.append(["MS002", "已完成", 25000, 0.05])
    ws1.append(["MS003", "已完成", 8000, 0.08])

    ws2 = wb.create_sheet("提成核算")
    ws2.append(["订单编号", "提成金额"])
    ws2.append(["MS001", 500])      # correct: 10000 * 0.05
    ws2.append(["MS002", 1200])     # wrong: should be 1250
    # MS003 missing

    wb.save(d / "combined.xlsx")
    wb.close()

    return d


class TestMultiSheetReconcile:
    """Group 5: multi-sheet reconciliation."""

    def test_parse_multi_sheet(self, multi_dir: Path) -> None:
        wb = parse_workbook(str(multi_dir / "combined.xlsx"))
        assert "销售明细" in wb.sheets
        assert "提成核算" in wb.sheets
        assert len(wb.sheets["销售明细"].rows) == 3
        assert len(wb.sheets["提成核算"].rows) == 2

    def test_reconcile_across_sheets(self, multi_dir: Path) -> None:
        wb = parse_workbook(str(multi_dir / "combined.xlsx"))
        result = reconcile(
            wb.sheets["销售明细"],
            wb.sheets["提成核算"],
            SheetMapping(source="销售明细", target="提成核算", key_column="订单编号"),
        )
        # MS001: match, MS002: no common numeric columns (source has 销售额/提成比例, target has 提成金额),
        #         so string compare finds no diff on key column
        # MS003: missing in target
        assert result.rows_checked == 3
        assert result.rows_mismatched >= 1
        # MS003 should be flagged as missing
        missing = [f for f in result.findings if "missing" in f.cause.lower()]
        assert len(missing) == 1
        assert missing[0].key_value == "MS003"

    def test_full_pipeline_multi_sheet(self, multi_dir: Path) -> None:
        dsl = RuleDsl(
            business_case="commission_check",
            inputs={"source_files": [str(multi_dir / "combined.xlsx")]},
            sheet_mapping=SheetMapping(source="销售明细", target="提成核算", key_column="订单编号"),
            rules=[
                ExcelAuditRule(id="r1", name="commission", when="订单状态 == 已完成", formula="销售额 * 提成比例"),
            ],
            outputs={},
        )
        output_dir = multi_dir / "artifacts"
        report = run_audit(dsl, job_id="multi_001", output_dir=output_dir)

        assert report.status == "completed"
        assert report.result.rows_checked > 0

    def test_report_generates_markdown(self, tmp_path: Path) -> None:
        report = AuditReport(
            job_id="test_report",
            status="completed",
            result=AuditResult(
                rules_evaluated=2,
                rows_checked=10,
                findings=[
                    AuditFinding(
                        row_index=0,
                        rule_id="r1",
                        severity=AuditSeverity.WARNING,
                        key_value="K1",
                        expected=100.0,
                        actual=90.0,
                        difference=10.0,
                        cause="Value mismatch",
                    ),
                ],
            ),
        )
        output = generate_report(report, tmp_path / "report_out")

        md_path = tmp_path / "report_out" / "report.md"
        assert md_path.exists()
        content = md_path.read_text()
        assert "Excel Audit Report" in content
        assert "10" in content  # rows checked

        json_path = tmp_path / "report_out" / "report.json"
        assert json_path.exists()

        diff_path = tmp_path / "report_out" / "diff.xlsx"
        assert diff_path.exists()

        assert len(output.artifacts) == 3

    def test_report_with_no_findings(self, tmp_path: Path) -> None:
        report = AuditReport(
            job_id="clean",
            status="completed",
            result=AuditResult(rules_evaluated=1, rows_checked=50),
        )
        output = generate_report(report, tmp_path / "clean_out")
        md = (tmp_path / "clean_out" / "report.md").read_text()
        assert "No mismatches" in md

    def test_report_diff_xlsx_has_findings(self, tmp_path: Path) -> None:
        from openpyxl import load_workbook as load

        report = AuditReport(
            job_id="diff_test",
            result=AuditResult(
                findings=[
                    AuditFinding(row_index=0, rule_id="r1", key_value="K", expected=1, actual=2, difference=1, cause="x"),
                ],
            ),
        )
        generate_report(report, tmp_path / "diff_out")

        wb = load(str(tmp_path / "diff_out" / "diff.xlsx"), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2  # header + 1 finding
        wb.close()
