"""Report generator for Excel audit results.

Produces three artifact types:
  - report.md: Human-readable markdown summary
  - report.json: Machine-readable JSON
  - diff.xlsx: Excel workbook with highlighted differences

No LLM — deterministic output from AuditResult.
"""
from __future__ import annotations

import json
import logging
from pathlib import Path

from excel_audit.contracts import AuditReport, AuditResult

logger = logging.getLogger(__name__)


def _build_markdown(report: AuditReport) -> str:
    """Generate a human-readable markdown report."""
    result = report.result
    lines: list[str] = [
        f"# Excel Audit Report",
        f"",
        f"**Job ID:** {report.job_id}",
        f"**Status:** {report.status}",
        f"",
        f"## Summary",
        f"",
        f"| Metric | Value |",
        f"|--------|-------|",
        f"| Rules evaluated | {result.rules_evaluated} |",
        f"| Rows checked | {result.rows_checked} |",
        f"| Rows mismatched | {result.rows_mismatched} |",
        f"| Mismatch amount total | {result.mismatch_amount_total:.2f} |",
        f"",
    ]

    if not result.findings:
        lines.append("✅ No mismatches found.")
        lines.append("")
    else:
        lines.append("## Row-Level Differences")
        lines.append("")
        lines.append("| Row | Rule | Key | Expected | Actual | Diff | Cause |")
        lines.append("|-----|------|-----|----------|--------|------|-------|")
        for f in result.findings:
            diff_str = f"{f.difference:.2f}" if f.difference is not None else "-"
            lines.append(
                f"| {f.row_index} | {f.rule_id} | {f.key_value} "
                f"| {f.expected} | {f.actual} | {diff_str} | {f.cause} |"
            )
        lines.append("")

        # Suspected causes
        causes = sorted({f.cause for f in result.findings if f.cause})
        if causes:
            lines.append("## Suspected Causes")
            lines.append("")
            for c in causes:
                lines.append(f"- {c}")
            lines.append("")

    lines.append(f"**Next action:** {report.next_action}")
    return "\n".join(lines)


def _build_diff_workbook(report: AuditReport, output_path: Path) -> None:
    """Generate an xlsx file with diff highlights."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        logger.warning("openpyxl not available, skipping diff.xlsx generation")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "diffs"

    # Headers
    headers = ["row_index", "rule_id", "severity", "key_value", "expected", "actual", "difference", "cause"]
    ws.append(headers)

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    for f in report.result.findings:
        row_data = [
            f.row_index, f.rule_id, f.severity.value, f.key_value,
            f.expected, f.actual, f.difference, f.cause,
        ]
        ws.append(row_data)
        fill = red_fill if f.severity.value == "error" else yellow_fill
        for col_idx in range(1, len(row_data) + 1):
            ws.cell(row=ws.max_row, column=col_idx).fill = fill

    wb.save(str(output_path))
    wb.close()


def generate_report(
    report: AuditReport,
    output_dir: str | Path,
) -> AuditReport:
    """Generate all report artifacts.

    Args:
        report: The audit report with results.
        output_dir: Directory to write artifacts to.

    Returns:
        Updated report with artifact paths filled in.
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    artifacts: list[str] = []

    # 1. report.md
    md_path = out / "report.md"
    md_path.write_text(_build_markdown(report), encoding="utf-8")
    artifacts.append(str(md_path))

    # 2. report.json
    json_path = out / "report.json"
    json_path.write_text(report.model_dump_json(indent=2), encoding="utf-8")
    artifacts.append(str(json_path))

    # 3. diff.xlsx
    diff_path = out / "diff.xlsx"
    _build_diff_workbook(report, diff_path)
    if diff_path.exists():
        artifacts.append(str(diff_path))

    return report.model_copy(update={"artifacts": artifacts})
