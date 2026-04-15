"""Excel workbook parser using openpyxl.

Reads .xlsx files into ParsedWorkbook / ParsedSheet structures.
No LLM involved — pure deterministic I/O.
"""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from excel_audit.contracts import ParsedSheet, ParsedWorkbook

logger = logging.getLogger(__name__)


def _coerce_cell(value: object) -> object:
    """Normalise cell values for downstream comparison."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def parse_workbook(
    file_path: str | Path,
    *,
    sheet_names: list[str] | None = None,
    header_row: int = 1,
    data_start_row: int | None = None,
) -> ParsedWorkbook:
    """Parse an .xlsx file into structured data.

    Args:
        file_path: Path to the .xlsx file.
        sheet_names: Optional list of sheet names to parse.
            If None, all sheets are parsed.
        header_row: 1-based row index for headers (default 1).
        data_start_row: 1-based row index where data begins.
            Defaults to header_row + 1.

    Returns:
        ParsedWorkbook with all requested sheets.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file format is invalid.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    if data_start_row is None:
        data_start_row = header_row + 1

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Failed to open workbook {path}: {exc}") from exc

    sheets: dict[str, ParsedSheet] = {}
    target_names = sheet_names or wb.sheetnames

    for name in target_names:
        if name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in %s, skipping", name, path.name)
            continue

        ws = wb[name]
        rows_iter = ws.iter_rows(min_row=header_row, values_only=True)

        # Extract headers
        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            sheets[name] = ParsedSheet(sheet_name=name)
            continue

        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]

        # Extract data rows
        data_rows: list[dict[str, object]] = []
        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if all(c is None for c in row):
                continue  # skip blank rows
            record = {headers[i]: _coerce_cell(cell) for i, cell in enumerate(row) if i < len(headers)}
            data_rows.append(record)

        sheets[name] = ParsedSheet(sheet_name=name, headers=headers, rows=data_rows)

    wb.close()
    return ParsedWorkbook(file_path=str(path), sheets=sheets)
